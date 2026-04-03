import sys
import json
import re
import xml.etree.ElementTree as ET
import zipfile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dataclasses import dataclass
from typing import List, Dict, Tuple, Any
from difflib import SequenceMatcher
import os

# --- CẤU HÌNH ---
SIMILARITY_THRESHOLD = 1  # Hạ xuống 0.70 để bù đắp các lỗi format công thức

@dataclass
class DocxQuestion:
    index: int
    question_text: str
    options: list
    answer: str

@dataclass
class SystemQuestion:
    index: int
    question_text: str
    options: list

@dataclass
class ExcelAnswers:
    tn: dict
    ds: dict
    dien: dict

@dataclass
class ErrorItem:
    error_type: str
    sub_type: str
    loai_cau: str
    stt: int
    detail_docx: str
    detail_system: str
    ghi_chu: str

# ==========================================
# CÁC HÀM XỬ LÝ TEXT & SIMILARITY
# ==========================================

def clean_html(text):
    """Làm sạch các tag HTML và Math để so sánh Text dễ dàng hơn."""
    if not text: return ""
    text = str(text)
    # Xóa các tag đặc biệt của công thức
    text = re.sub(r'<m:[^>]+>', '', text)
    # Xóa HTML
    text = re.sub(r'<[^>]+>', '', text)
    # Decode 
    text = text.replace('&nbsp;', ' ').replace('&lt;', '<').replace('&gt;', '>')
    # Xóa multiple spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def text_similarity(t1, t2):
    """Tính tỷ lệ giống nhau giữa 2 đoạn text."""
    if not t1 or not t2: return 0
    s1 = clean_html(t1).lower()
    s2 = clean_html(t2).lower()
    return SequenceMatcher(None, s1, s2).ratio()

def find_best_match(target_text, candidate_list):
    """Tìm câu hỏi có nội dung giống nhất trong danh sách JSON HT."""
    best_match = None
    best_score = 0
    for cand in candidate_list:
        cand_text = cand.question_text if hasattr(cand, 'question_text') else cand
        score = text_similarity(target_text, cand_text)
        if score > best_score:
            best_score = score
            best_match = cand
    return best_match, best_score

def find_best_match_idx(target_text, options_list):
    """Tìm vị trí option (A,B,C,D) chứa text giống nhất."""
    best_idx = None
    best_score = 0
    for idx, opt_text in enumerate(options_list):
        score = text_similarity(target_text, opt_text)
        if score > best_score:
            best_score = score
            best_idx = idx
    return best_idx, best_score

# ==========================================
# HÀM LOAD AI ANSWERS
# ==========================================

def load_ai_pdf_answers(json_path):
    """Đọc và tổ chức lại kết quả AI bóc tách từ PDF theo cấu trúc chuẩn."""
    with open(json_path, 'r', encoding='utf-8') as f:
        raw_data = json.load(f)
        
    ai_ans = {"TN": [], "DS": [], "DIEN": []}
    for item in raw_data:
        sec = item.get("loai_cau", "")
        if sec in ai_ans:
            ai_ans[sec].append(item)
            
    # Lấy số thứ tự từ chuỗi (tránh lỗi font/text)
    def extract_num(s):
        m = re.search(r'\d+', str(s))
        return int(m.group()) if m else 0
        
    # Sort từng phần theo câu số
    for sec in ai_ans:
        ai_ans[sec].sort(key=lambda x: extract_num(x.get("cau_so", 0)))
        
    # Gom nhóm phần Đúng Sai (Vì 1 câu có 4 ý a,b,c,d)
    ds_grouped = {}
    for item in ai_ans["DS"]:
        c_so = extract_num(item.get("cau_so", 0))
        if c_so not in ds_grouped:
            ds_grouped[c_so] = {}
        y = item.get("y_phu", "").lower()
        if y:
            ds_grouped[c_so][y] = item
            
    ds_list = []
    for c_so in sorted(ds_grouped.keys()):
        ds_list.append(ds_grouped[c_so])
        
    ai_ans["DS"] = ds_list
    return ai_ans

# ==========================================
# CÁC HÀM ĐỌC DOCX, JSON, EXCEL (GIỮ NGUYÊN TỪ CODE CŨ)
# ==========================================
from docx import Document
from docx.oxml.ns import qn

def extract_full_text_including_math(para):
    """Hàm đọc sâu vào XML để lấy cả text thường (<w:t>) và text công thức toán (<m:t>)."""
    text_parts = []
    for node in para._element.iter():
        if node.tag == qn('w:t') and node.text:
            text_parts.append(node.text)
        elif node.tag == qn('m:t') and node.text:
            text_parts.append(node.text)
    return "".join(text_parts).strip()

def parse_docx_data(docx_path: str) -> Dict[str, List[DocxQuestion]]:
    doc = Document(docx_path)
    data = {"TN": [], "DS": [], "DIEN": []}
    current_sec = None
    q_index = 0
    
    current_q_text = ""
    current_opts = []
    
    def save_q():
        nonlocal current_q_text, current_opts
        if current_sec and current_q_text:
            data[current_sec].append(DocxQuestion(
                index=q_index,
                question_text=current_q_text.strip(),
                options=current_opts,
                answer="" # Bypass answer, dùng AI sau
            ))
        current_q_text = ""
        current_opts = []

    for para in doc.paragraphs:
        text = extract_full_text_including_math(para)
        if not text: continue
        
        t_upper = text.upper()
        if "PHẦN 1" in t_upper or "TRẮC NGHIỆM NHIỀU PHƯƠNG ÁN" in t_upper:
            save_q()
            current_sec = "TN"
            q_index = 0
            continue
        elif "PHẦN 2" in t_upper or "TRẮC NGHIỆM ĐÚNG SAI" in t_upper:
            save_q()
            current_sec = "DS"
            q_index = 0
            continue
        elif "PHẦN 3" in t_upper or "TRẢ LỜI NGẮN" in t_upper:
            save_q()
            current_sec = "DIEN"
            q_index = 0
            continue
            
        if re.match(r'^Câu\s+\d+[:.]', text, re.I):
            save_q()
            q_index += 1
            current_q_text = text
        elif re.match(r'^[A-D]\.', text):
            current_opts.append(text)
        else:
            if current_q_text and not current_opts:
                current_q_text += "\n" + text
            elif current_opts:
                current_opts[-1] += "\n" + text

    save_q()
    return data

def parse_excel_answers(excel_path: str) -> ExcelAnswers:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    ans_tn = {}
    ans_ds = {}
    ans_dien = {}
    
    current_section = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        
        part_str = str(row[0]).strip().lower()
        if "phần 1" in part_str: current_section = 1
        elif "phần 2" in part_str: current_section = 2
        elif "phần 3" in part_str: current_section = 3
        else:
            stt = int(row[0]) if str(row[0]).isdigit() else 0
            if stt == 0: continue
            
            if current_section == 1:
                ans_tn[stt] = str(row[1]).strip() if row[1] else ""
            elif current_section == 2:
                ans_ds[stt] = {
                    'a': str(row[1]).strip() if row[1] else "",
                    'b': str(row[2]).strip() if row[2] else "",
                    'c': str(row[3]).strip() if row[3] else "",
                    'd': str(row[4]).strip() if row[4] else ""
                }
            elif current_section == 3:
                ans_dien[stt] = str(row[1]).strip() if row[1] else ""

    return ExcelAnswers(tn=ans_tn, ds=ans_ds, dien=ans_dien)

def parse_system_json(json_path: str) -> Dict[str, List[SystemQuestion]]:
    with open(json_path, 'r', encoding='utf-8') as f:
        jdata = json.load(f)
    
    sys_data = {"TN": [], "DS": [], "DIEN": []}
    
    for section in jdata.get("data", []):
        sec_idx = section.get("indexSection")
        sec_key = "TN" if sec_idx == 1 else "DS" if sec_idx == 2 else "DIEN" if sec_idx == 3 else None
        if not sec_key: continue
        
        for q in section.get("dataQuestion", []):
            q_idx = q.get("indexQuestion", 0)
            q_text = q.get("contentQuestion", "")
            opts = [opt.get("contentOption", "") for opt in q.get("options", [])]
            
            sys_data[sec_key].append(SystemQuestion(
                index=q_idx,
                question_text=q_text,
                options=opts
            ))
    return sys_data

# ==========================================
# CORE: SO SÁNH PHÁT HIỆN LỖI TRỘN ĐỀ
# ==========================================

def detect_all_errors(docx_data, excel_answers, json_data, ai_answers):
    loi_dap_an = []
    loi_thua_nd = []
    loi_sai_nd = []

    CFGS = [
        ("TN", "Trắc nghiệm", excel_answers.tn),
        ("DS", "Đúng sai",    excel_answers.ds),
        ("DIEN", "Điền",      excel_answers.dien),
    ]

    for sec_key, sec_name, excel_dict in CFGS:
        dq_list = docx_data.get(sec_key, [])
        jq_list = json_data.get(sec_key, [])
        ai_list = ai_answers.get(sec_key, [])

        matched_json_indices = set()

        for i, d in enumerate(dq_list):
            ai_ans = ai_list[i] if i < len(ai_list) else None
            
            # 1. Tìm bản sao của Câu hỏi này bên Hệ thống (Chống trộn câu hỏi)
            j, score = find_best_match(d.question_text, jq_list)
            
            if not j or score < SIMILARITY_THRESHOLD:
                loi_sai_nd.append(ErrorItem(
                    error_type="Lỗi sai nội dung", sub_type="Thiếu câu trên HT",
                    loai_cau=sec_name, stt=d.index,
                    detail_docx=f"Câu {d.index}: {d.question_text[:80]}...", 
                    detail_system="(Không tìm thấy câu tương đương trên HT)",
                    ghi_chu=f"Không có câu nào trên JSON hệ thống khớp với nội dung gốc."
                ))
                continue
            
            matched_json_indices.add(j.index)
            sys_ans_excel = excel_dict.get(j.index) # Đáp án chuẩn của HT cho câu bị trộn này

            # 2. Kiểm tra đáp án có bị sai do trộn không
            if sec_key == "TN":
                if ai_ans:
                    pdf_correct_text = clean_html(str(ai_ans.get("noi_dung", "")))
                    pdf_dap_an_char = str(ai_ans.get("dap_an", "")).strip().upper()
                    
                    sys_correct_text = ""
                    sys_opt_char = str(sys_ans_excel).strip().upper()
                    if sys_opt_char in ['A', 'B', 'C', 'D']:
                        idx = ord(sys_opt_char) - ord('A')
                        if idx < len(j.options):
                            sys_correct_text = clean_html(j.options[idx])
                    
                    score_match = text_similarity(pdf_correct_text, sys_correct_text)
                    if score_match < SIMILARITY_THRESHOLD:
                        loi_dap_an.append(ErrorItem(
                            error_type="Lỗi đáp án", sub_type="Sai đáp án (Do trộn Option)",
                            loai_cau=sec_name, stt=d.index,
                            detail_docx=f"{pdf_dap_an_char}: {pdf_correct_text[:80]}",
                            detail_system=f"HT chọn {sys_opt_char}: {sys_correct_text[:80]}",
                            ghi_chu=f"Map với câu {j.index} HT. Đáp án HT trỏ sai nội dung (Độ khớp: {score_match*100:.0f}%)."
                        ))

            elif sec_key == "DS":
                if ai_ans:
                    for y_phu in ['a', 'b', 'c', 'd']:
                        pdf_sub = ai_ans.get(y_phu)
                        if not pdf_sub: continue
                        
                        pdf_sub_text = clean_html(str(pdf_sub.get("noi_dung", "")))
                        pdf_sub_tf = str(pdf_sub.get("dap_an", "")).strip().upper()
                        
                        # Quét tìm xem ý phụ này đang nằm ở đâu bên hệ thống (bị trộn)
                        sys_opt_idx, opt_score = find_best_match_idx(pdf_sub_text, j.options)
                        
                        if sys_opt_idx is not None and opt_score > SIMILARITY_THRESHOLD:
                            sys_y_phu_char = chr(ord('a') + sys_opt_idx)
                            sys_sub_tf = str(sys_ans_excel.get(sys_y_phu_char, "")).strip().upper()
                            
                            # Chuẩn hóa TRUE/FALSE sang Đ/S
                            sys_sub_tf = 'Đ' if sys_sub_tf == 'TRUE' else 'S' if sys_sub_tf == 'FALSE' else sys_sub_tf
                            pdf_sub_tf = 'Đ' if pdf_sub_tf == 'TRUE' else 'S' if pdf_sub_tf == 'FALSE' else pdf_sub_tf

                            if sys_sub_tf and pdf_sub_tf != sys_sub_tf:
                                sys_opt_text = clean_html(j.options[sys_opt_idx])
                                loi_dap_an.append(ErrorItem(
                                    error_type="Lỗi đáp án", sub_type="Sai ý Đúng/Sai",
                                    loai_cau=sec_name, stt=d.index,
                                    detail_docx=f"Ý '{y_phu}' PDF: [{pdf_sub_tf}] - {pdf_sub_text[:50]}",
                                    detail_system=f"Map vào ý '{sys_y_phu_char}' HT: [{sys_sub_tf}] - {sys_opt_text[:50]}",
                                    ghi_chu=f"Cùng một nội dung nhưng Hệ Thống chấm sai (Đ/S bị lệch)."
                                ))
                        else:
                            loi_sai_nd.append(ErrorItem(
                                error_type="Lỗi sai nội dung", sub_type="Mất ý Đúng/Sai",
                                loai_cau=sec_name, stt=d.index,
                                detail_docx=f"Ý '{y_phu}': {pdf_sub_text[:60]}",
                                detail_system="(Không tìm thấy ý tương đương trên Hệ thống)",
                                ghi_chu=f"Không map được ý '{y_phu}' của DOCX sang JSON."
                            ))

            elif sec_key == "DIEN":
                if ai_ans:
                    pdf_dap_an = str(ai_ans.get("dap_an", "")).strip()
                    sys_dap_an = str(sys_ans_excel).strip()
                    
                    # Cân bằng dấu phẩy và dấu chấm
                    pdf_norm = pdf_dap_an.replace(',', '.')
                    sys_norm = sys_dap_an.replace(',', '.')
                    
                    if pdf_norm != sys_norm:
                        loi_dap_an.append(ErrorItem(
                            error_type="Lỗi đáp án", sub_type="Sai số liệu Điền khuyết",
                            loai_cau=sec_name, stt=d.index,
                            detail_docx=f"PDF bóc được: {pdf_dap_an}",
                            detail_system=f"Hệ thống nhập: {sys_dap_an}",
                            ghi_chu=f"Câu {d.index} DOCX map với câu {j.index} HT. Đáp án không trùng khớp."
                        ))

        # 3. Lỗi thừa nội dung (Có trên JSON HT mà DOCX không có)
        for j in jq_list:
            if j.index not in matched_json_indices:
                loi_thua_nd.append(ErrorItem(
                    error_type="Lỗi thừa nội dung", sub_type="Thừa câu trên HT",
                    loai_cau=sec_name, stt=j.index, 
                    detail_docx="(Không có trong gốc)", detail_system=(clean_html(j.question_text)[:100] + "..."),
                    ghi_chu=f"Câu {j.index} trên HT không match với bất kỳ câu nào trong đề DOCX gốc."
                ))

    return loi_dap_an, loi_thua_nd, loi_sai_nd

# ==========================================
# XUẤT BÁO CÁO EXCEL (GIỮ NGUYÊN)
# ==========================================

COLOR = {"Trắc nghiệm": "E8F5E9", "Đúng sai": "FFF3E0", "Điền": "E3F2FD"}

def _thin_border():
    thin = Side(border_style="thin", color="CBD5E1")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, name="Arial", size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="3B82F6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _thin_border()

def _write_error_sheet(wb, sheet_name, error_list, title, desc):
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = f"{title.upper()} - {desc}"
    t.font = Font(bold=True, name="Arial", size=14, color="1E3A5F")
    t.fill = PatternFill("solid", fgColor="EFF6FF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Loại câu", "STT (Gốc)", "Sub-type", "Nội dung DOCX/PDF", "Nội dung Hệ thống", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 2, col, h)
    ws.row_dimensions[2].height = 25

    r = 3
    for e in error_list:
        bg = COLOR.get(e.loai_cau, "FFFFFF")
        data = [
            (e.loai_cau, bg, False, "1E3A5F"),
            (e.stt, bg, True, "1E3A5F"),
            (e.sub_type, bg, False, "B91C1C"),
            (e.detail_docx, "F9FAFB", False, "374151"),
            (e.detail_system, "FEFCE8", False, "374151"),
            (e.ghi_chu, "FFFFFF", True, "047857")
        ]

        for col, (val, cell_bg, bold, text_color) in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(bold=bold, name="Arial", size=10, color=text_color)
            c.fill = PatternFill("solid", fgColor=cell_bg)
            c.alignment = Alignment(horizontal="center" if col in [1, 2, 3] else "left", vertical="center", wrap_text=True)
            c.border = _thin_border()
        
        ws.row_dimensions[r].height = 45
        r += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 40

def _write_summary(wb, docx_data, json_data, excel_ans, loi_da, loi_thua, loi_sai):
    ws = wb.active
    ws.title = "TỔNG QUAN"
    
    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "BÁO CÁO KIỂM TRA LỖI TRỘN ĐỀ HT"
    t.font = Font(bold=True, name="Arial", size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1E3A8A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Bỏ qua phần code ghi thông tin chi tiết (Giữ ngắn gọn file cho bạn)
    r = 3
    ws.cell(row=r, column=1, value="Tổng lỗi đáp án (Do trộn/Sai):").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(loi_da)).font = Font(color="B91C1C", bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Tổng lỗi thừa câu:").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(loi_thua)).font = Font(color="B91C1C", bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Tổng lỗi sai/thiếu nội dung:").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(loi_sai)).font = Font(color="B91C1C", bold=True)
    
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

def export_report(docx_data, json_data, excel_ans, loi_da, loi_thua, loi_sai, out_path):
    wb = Workbook()
    _write_summary(wb, docx_data, json_data, excel_ans, loi_da, loi_thua, loi_sai)
    _write_error_sheet(wb, "1. Lỗi đáp án", loi_da, "Lỗi đáp án", "Sai lệch khi Map chéo nội dung (Chống trộn)")
    _write_error_sheet(wb, "2. Lỗi thừa nội dung", loi_thua, "Lỗi thừa nội dung", "Câu có trên HT nhưng không có trong gốc")
    _write_error_sheet(wb, "3. Lỗi sai nội dung", loi_sai, "Lỗi sai nội dung", "Mất câu hoặc mất ý phụ")
    
    wb.save(str(out_path))
    print(f"\n[+] ĐÃ XUẤT BÁO CÁO THÀNH CÔNG: {out_path}")

# ==========================================
# MAIN EXECUTION
# ==========================================

def main():
    # 1. Khai báo các file đầu vào
    docx_file = "D:\\CheckTool\\FetchJson\\Toán ĐỀ ONLINE 10 mới 2.docx"
    json_file = "D:\\CheckTool\\FetchJson\\onluyen_data\\69cc7f41b86b57e0fac47636.json"
    excel_file = "D:\\CheckTool\\FetchJson\\onluyen_data\\69cc7f41b86b57e0fac47636.xlsx" # Hoặc file .xlsx
    ai_answers_file = "D:\\CheckTool\\FetchJson\\dapan_tu_pdf.json"
    out_file = "BAO_CAO_CHECK_DE.xlsx"

    print("="*50)
    print("🚀 KHỞI ĐỘNG HỆ THỐNG SOÁT LỖI CHỐNG TRỘN ĐỀ")
    print("="*50)

    if not os.path.exists(ai_answers_file):
        print(f"[!] THIẾU FILE: Không tìm thấy {ai_answers_file}")
        print("[!] Hãy chạy script AI bóc tách PDF trước để tạo file này.")
        return

    print("1. Đang load đáp án chuẩn từ AI (dapan_tu_pdf.json)...")
    ai_answers = load_ai_pdf_answers(ai_answers_file)

    print(f"2. Đang đọc câu hỏi gốc từ DOCX: {docx_file}...")
    docx_data = parse_docx_data(docx_file)

    print(f"3. Đang đọc cấu trúc từ hệ thống: {json_file}...")
    json_data = parse_system_json(json_file)

    print(f"4. Đang đọc đáp án hệ thống: {excel_file}...")
    excel_ans = parse_excel_answers(excel_file)

    print("5. Đang chạy thuật toán Khớp Nối Ngữ Nghĩa (Content Mapping) & Check Lỗi...")
    loi_da, loi_thua, loi_sai = detect_all_errors(docx_data, excel_ans, json_data, ai_answers)

    export_report(docx_data, json_data, excel_ans, loi_da, loi_thua, loi_sai, out_file)

if __name__ == "__main__":
    main()