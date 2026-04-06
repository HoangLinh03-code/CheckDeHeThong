import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def export_to_excel(results_data, summary_data, excel_path):
    # Đảm bảo thư mục export tồn tại
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: KẾT QUẢ CHI TIẾT
    # ---------------------------------------------------------
    ws_results = wb.active
    ws_results.title = "Kết Quả Chi Tiết"
    
    # Cấu hình Style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border_style = Side(style='thin', color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    headers = ["STT", "Câu hỏi đề gốc", "Câu hỏi đề hệ thống", "Trạng thái / Lỗi"]
    ws_results.append(headers)
    
    # Trang trí Header
    for col in range(1, 5):
        cell = ws_results.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
        
    # Ghi dữ liệu
    for row_idx, row in enumerate(results_data, start=2):
        ws_results.cell(row=row_idx, column=1, value=row["STT"]).alignment = align_center
        ws_results.cell(row=row_idx, column=2, value=row["Câu hỏi đề gốc"]).alignment = align_center
        ws_results.cell(row=row_idx, column=3, value=row["Câu hỏi đề hệ thống"]).alignment = align_center
        
        error_cell = ws_results.cell(row=row_idx, column=4, value=row["Lỗi"])
        if "✅" in row["Lỗi"]:
            error_cell.alignment = align_center
            error_cell.font = Font(color="008000", bold=True) # Xanh lá
        else:
            error_cell.alignment = align_left
            error_cell.font = Font(color="C00000", bold=True) # Đỏ đậm
            
        for col in range(1, 5):
            ws_results.cell(row=row_idx, column=col).border = border
            
    # Chỉnh độ rộng cột
    ws_results.column_dimensions['A'].width = 8
    ws_results.column_dimensions['B'].width = 25
    ws_results.column_dimensions['C'].width = 25
    ws_results.column_dimensions['D'].width = 45

    # ---------------------------------------------------------
    # SHEET 2: THỐNG KÊ TỔNG KẾT
    # ---------------------------------------------------------
    ws_summary = wb.create_sheet(title="Thống Kê Tổng Kết")
    summary_headers = ["Chỉ số", "Số lượng"]
    ws_summary.append(summary_headers)
    
    sum_header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    for col in range(1, 3):
        cell = ws_summary.cell(row=1, column=col)
        cell.fill = sum_header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
        
    for row_idx, (key, val) in enumerate(summary_data.items(), start=2):
        ws_summary.cell(row=row_idx, column=1, value=key).alignment = align_left
        cell_val = ws_summary.cell(row=row_idx, column=2, value=val)
        cell_val.alignment = align_center
        cell_val.font = Font(bold=True)
        
        for col in range(1, 3):
            ws_summary.cell(row=row_idx, column=col).border = border
            
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 20

    wb.save(excel_path)