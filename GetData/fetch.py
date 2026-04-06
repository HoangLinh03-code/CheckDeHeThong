"""
fetch_and_export.py
Fetch JSON tu onluyen.vn -> Parse cau hoi -> Export Excel

Cai thu vien: pip install requests openpyxl

Cach dung:
  # Fetch tu URL preview
  python fetch.py --url "https://app.onluyen.vn/preview-exam?url=..."

  # Dung file JSON da tai san
  python fetch.py --file onluyen_data/abc123.json
"""

import argparse
import json
import re
import sys
from itertools import groupby
from pathlib import Path
from urllib.parse import unquote, urlparse, parse_qs

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ======================================================================
# CONSTANTS
# ======================================================================

OUTPUT_DIR = Path("onluyen_data")
OUTPUT_DIR.mkdir(exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/123.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, */*",
    "Accept-Language": "vi-VN,vi;q=0.9",
    "Referer": "https://app.onluyen.vn/",
    "Origin": "https://app.onluyen.vn",
}

# typeAnswer -> ten dang cau
TYPE_ANSWER_LABEL = {
    0: "Trac nghiem",
    1: "Dung sai",
    5: "Dien",
}

# idOption (0,1,2,3) -> chu cai (A,B,C,D)
OPTION_LABELS = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E"}

# true/false -> D/S
TRUE_FALSE_MAP = {"true": "\u0110", "false": "S"}

# Mau nen theo dang cau
COLOR_ROW = {
    "Trac nghiem": "DBEAFE",
    "Dung sai":    "FEF9C3",
    "Dien":        "DCFCE7",
}


# ======================================================================
# UTILS
# ======================================================================

def strip_html(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", "", text)
    for entity, char in [("&nbsp;", " "), ("&lt;", "<"), ("&gt;", ">"), ("&amp;", "&")]:
        text = text.replace(entity, char)
    return re.sub(r"\s+", " ", text).strip()


def extract_json_url(preview_url: str) -> str:
    """Lay URL JSON thuc tu preview page URL."""
    parsed = urlparse(preview_url)
    params = parse_qs(parsed.query)
    raw = params.get("url", [None])[0]
    if not raw:
        raise ValueError("Khong tim thay param 'url' trong preview URL.")
    return unquote(raw)


def json_id_from_url(json_url: str) -> str:
    """Lay phan ten file (khong co .json) lam ID."""
    return Path(urlparse(json_url).path).stem


# ======================================================================
# STEP 1: LOAD DATA (tu URL hoac file)
# ======================================================================

def load_from_url(preview_url: str) -> tuple[dict, str]:
    """Fetch JSON tu preview URL. Tra ve (data, json_id)."""
    # Preview URL hay JSON URL deu duoc
    if "preview-exam" in preview_url:
        json_url = extract_json_url(preview_url)
    else:
        json_url = preview_url

    session = requests.Session()
    session.headers.update(HEADERS)

    print(f"[->] Fetching: {json_url}")
    resp = session.get(json_url, timeout=15)
    resp.raise_for_status()

    data = resp.json()
    json_id = json_id_from_url(json_url)

    # Luu raw JSON
    raw_path = OUTPUT_DIR / f"{json_id}.json"
    raw_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] Da luu raw JSON: {raw_path}")

    return data, json_id


def load_from_file(file_path: str) -> tuple[dict, str]:
    """Doc JSON tu file local. Tra ve (data, json_id)."""
    path = Path(file_path)
    if not path.exists():
        print(f"[X] Khong tim thay file: {path}")
        sys.exit(1)

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    json_id = path.stem  # ten file khong co extension
    print(f"[OK] Da doc file: {path} (id={json_id})")
    return data, json_id


# ======================================================================
# STEP 2: PARSE CAU HOI
# ======================================================================

def format_answer(ds: dict, type_answer: int) -> str:
    """
    Tra ve dap an dung theo tung dang cau:
      - Trac nghiem (typeAnswer=0): answerOptionId[0] -> A/B/C/D
      - Dung sai    (typeAnswer=1): answerFreeText[]  -> D/S chuoi
      - Dien        (typeAnswer=5): answerFreeText[0] -> so/text
    """
    if type_answer == 0:
        ids = ds.get("answerOptionId", [])
        return OPTION_LABELS.get(ids[0], "?") if ids else "?"

    elif type_answer == 1:
        texts = ds.get("answerFreeText", [])
        return "".join(TRUE_FALSE_MAP.get(t.strip().lower(), "?") for t in texts)

    elif type_answer == 5:
        texts = ds.get("answerFreeText", [])
        return texts[0].strip() if texts else "?"

    return "?"


def parse_questions(data: dict) -> list[dict]:
    """
    Parse toan bo blocks thanh list cau hoi chuan hoa.

    Co 2 loai block:
      typeData=0 -> cau don: cau hoi nam trong block["dataStandard"]
      typeData=1 -> nhom cau (reading/passage): cau hoi nam trong
                    block["dataMaterial"]["data"] la 1 list
    STT cau dua theo stepIndex (0-based) + 1 de dam bao dung thu tu de thi.
    """
    questions = []

    for block in data.get("data", []):
        type_data = block.get("typeData")

        if type_data == 0:
            # Cau don
            ds          = block.get("dataStandard", {})
            type_answer = ds.get("typeAnswer", 0)
            stt         = ds.get("stepIndex", len(questions)) + 1
            questions.append({
                "stt":         stt,
                "dang":        TYPE_ANSWER_LABEL.get(type_answer, f"type={type_answer}"),
                "dap_an":      format_answer(ds, type_answer),
                "type_answer": type_answer,
                "max_score":   ds.get("maxScore", 0),
            })

        elif type_data == 1:
            # Nhom cau co doan van (reading passage)
            sub_list = block.get("dataMaterial", {}).get("data", [])
            for ds in sub_list:
                type_answer = ds.get("typeAnswer", 0)
                stt         = ds.get("stepIndex", len(questions)) + 1
                questions.append({
                    "stt":         stt,
                    "dang":        TYPE_ANSWER_LABEL.get(type_answer, f"type={type_answer}"),
                    "dap_an":      format_answer(ds, type_answer),
                    "type_answer": type_answer,
                    "max_score":   ds.get("maxScore", 0),
                })

    # Sap xep lai theo STT de dam bao thu tu chinh xac
    questions.sort(key=lambda q: q["stt"])
    return questions


# ======================================================================
# STEP 3: EXPORT EXCEL
# ======================================================================

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_excel(questions: list[dict], exam_info: dict, out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dap an"

    # Row 1: Tieu de de thi
    ws.merge_cells("A1:C1")
    title            = ws["A1"]
    title.value      = f"DAP AN: {exam_info.get('name', '').upper()}"
    title.font       = Font(bold=True, name="Arial", size=14, color="1E3A5F")
    title.alignment  = Alignment(horizontal="center", vertical="center")
    title.fill       = PatternFill("solid", fgColor="EFF6FF")
    ws.row_dimensions[1].height = 36

    # Row 2: Header
    hfill = PatternFill("solid", fgColor="1E3A5F")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(["Dang", "STT Cau", "Dap an"], 1):
        cell            = ws.cell(row=2, column=col, value=h)
        cell.fill       = hfill
        cell.font       = hfont
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = _border()
    ws.row_dimensions[2].height = 28

    # Rows 3+: Du lieu, merge cot Dang theo nhom
    current_row = 3
    for dang, group in groupby(questions, key=lambda q: q["dang"]):
        group_list = list(group)
        fill       = PatternFill("solid", fgColor=COLOR_ROW.get(dang, "FFFFFF"))
        start_row  = current_row

        for q in group_list:
            # Xử lý format cho cột A (Dạng) trên từng dòng để giữ nguyên Border/Fill khi Merge
            cell_dang = ws.cell(row=current_row, column=1)
            if current_row == start_row:
                cell_dang.value = dang  # Chỉ gán text ở dòng đầu tiên của nhóm
            cell_dang.fill       = fill
            cell_dang.border     = _border()
            cell_dang.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_dang.font       = Font(bold=True, name="Arial", size=10)

            # Xử lý cột B (STT)
            cell_stt            = ws.cell(row=current_row, column=2, value=q["stt"])
            cell_stt.fill       = fill
            cell_stt.border     = _border()
            cell_stt.alignment  = Alignment(horizontal="center", vertical="center")
            cell_stt.font       = Font(name="Arial", size=10)

            # Xử lý cột C (Đáp án)
            cell_ans            = ws.cell(row=current_row, column=3, value=q["dap_an"])
            cell_ans.fill       = fill
            cell_ans.border     = _border()
            cell_ans.alignment  = Alignment(horizontal="center", vertical="center")
            cell_ans.font       = Font(bold=True, name="Arial", size=12, color="C0392B")

            ws.row_dimensions[current_row].height = 28
            current_row += 1

        # Merge cells cột A theo dạng
        end_row = current_row - 1
        if start_row < end_row:
            ws.merge_cells(f"A{start_row}:A{end_row}")

        cell_dang            = ws.cell(row=start_row, column=1, value=dang)
        cell_dang.fill       = fill
        cell_dang.border     = _border()
        cell_dang.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_dang.font       = Font(bold=True, name="Arial", size=10)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A3"

    # Sheet 2: Thong ke
    ws2 = wb.create_sheet("Thong ke")
    for col, h in enumerate(["Dang cau", "So luong", "Tong diem"], 1):
        cell            = ws2.cell(row=1, column=col, value=h)
        cell.font       = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill       = PatternFill("solid", fgColor="1E3A5F")
        cell.border     = _border()

    stats: dict[str, dict] = {}
    for q in questions:
        d = q["dang"]
        if d not in stats:
            stats[d] = {"count": 0, "score": 0.0}
        stats[d]["count"] += 1
        stats[d]["score"] += q["max_score"]

    for row_idx, (dang, s) in enumerate(stats.items(), 2):
        fill = PatternFill("solid", fgColor=COLOR_ROW.get(dang, "FFFFFF"))
        for col, val in enumerate([dang, s["count"], round(s["score"], 4)], 1):
            cell        = ws2.cell(row=row_idx, column=col, value=val)
            cell.fill   = fill
            cell.border = _border()
            cell.font   = Font(name="Arial", size=10)

    total_row = len(stats) + 2
    ws2.cell(row=total_row, column=1, value="TONG").font = Font(bold=True, name="Arial")
    ws2.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
    ws2.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 18

    wb.save(out_path)
    print(f"[OK] Da xuat Excel: {out_path}")


# ======================================================================
# MAIN
# ======================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch de thi tu onluyen.vn va export ra Excel"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--url",
        metavar="PREVIEW_URL",
        help='URL preview hoac URL JSON truc tiep. Vi du: --url "https://app.onluyen.vn/preview-exam?url=..."',
    )
    group.add_argument(
        "--file",
        metavar="JSON_FILE",
        help="Duong dan toi file JSON da tai ve. Vi du: --file onluyen_data/abc123.json",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # B1: Load du lieu
    if args.url:
        try:
            data, json_id = load_from_url(args.url)
        except requests.HTTPError as e:
            print(f"[X] HTTP Error {e.response.status_code}: {e}")
            sys.exit(1)
        except Exception as e:
            print(f"[X] Loi fetch: {e}")
            sys.exit(1)
    else:
        data, json_id = load_from_file(args.file)

    print(f"[->] De thi: '{data.get('name')}' | Tong cau: {data.get('totalQuestion')}")

    # B2: Parse cau hoi
    questions = parse_questions(data)
    print(f"[OK] Parse xong: {len(questions)} cau")
    for dang in TYPE_ANSWER_LABEL.values():
        count = sum(1 for q in questions if q["dang"] == dang)
        if count:
            print(f"     {dang}: {count} cau")

    # B3: Export Excel — ten file = json_id de tranh ghi de
    out_path = OUTPUT_DIR / f"{json_id}.xlsx"
    exam_info = {
        "name":          data.get("name", ""),
        "duration":      data.get("duration", 0),
        "totalQuestion": data.get("totalQuestion", 0),
        "maxScore":      data.get("maxScore", 0),
    }
    export_excel(questions, exam_info, out_path)


if __name__ == "__main__":
    main()